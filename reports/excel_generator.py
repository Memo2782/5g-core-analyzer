import pandas as pd
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class CoreExcelGenerator:
    def __init__(self):
        self.ruta_salida = "Reporte_Auditoria_5GC_Premium.xlsx"

    def generar_reporte(self, alertas):
        if not alertas:
            print("[-] No hay alertas para exportar al reporte Excel.")
            return None

        print(f"[+] Generando reporte corporativo avanzado en Excel: {self.ruta_salida}...")
        
        # 1. Map data directly with explicit presentation keys to avoid length mismatches
        datos_formateados = []
        for a in alertas:
            datos_formateados.append({
                'Estampa de Tiempo': a.get('timestamp', 'Desconocido'),
                'Dominio Procedimiento': a.get('procedimiento', '5GS_SIGNALING'),
                'Interfaz Core': a.get('interfaz', 'SBI'),
                'Nodo Origen': a.get('origen', 'Desconocido'),
                'Nodo Destino': a.get('destino', 'Desconocido'),
                'Código de Control': a.get('codigo', 'Desconocido'),
                'Error Oficial 3GPP': a.get('error_3gpp', 'Protocol Error'),
                'Causa Raíz Operativa': a.get('causa_raiz', 'Analizar traza de señalización profunda.'),
                'Acción Sugerida (Mantenimiento L3)': a.get('solucion', 'Revisar logs internos de la NF.')
            })
            
        df = pd.DataFrame(datos_formateados)

        try:
            with pd.ExcelWriter(self.ruta_salida, engine='openpyxl') as writer:
                # Inject dataframe onto sheet starting at row 5
                df.to_excel(writer, sheet_name='Análisis E2E', startrow=4, index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Análisis E2E']
                
                # --- PREMIUM UI LAYOUT STYLES ---
                fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
                fill_kpi_label = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
                fill_error_critico = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                
                font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
                font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                font_body = Font(name="Segoe UI", size=10, color="333333")
                font_kpi_val = Font(name="Segoe UI", size=12, bold=True, color="C00000")
                
                border_thin = Border(
                    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
                )
                
                # 2. Add Top Dashboard Elements
                worksheet['A1'] = "📶 COOPERATIVE MULTI-TRACE NETWORK AUDIT REPORT"
                worksheet['A1'].font = font_title
                
                worksheet['A3'] = "MÉTRICA CRÍTICA:"
                worksheet['B3'] = f"Total Anomalías E2E: {len(df)}"
                worksheet['A3'].font = Font(name="Segoe UI", size=10, bold=True, color="555555")
                worksheet['B3'].font = font_kpi_val
                worksheet['B3'].fill = fill_kpi_label
                worksheet['B3'].alignment = Alignment(horizontal="center")
                
                # 3. Format Headers (Row 5)
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=5, column=col_num)
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border_thin
                
                # 4. Format Data Rows & Apply soft red alerting to failures
                for row_idx in range(6, 6 + len(df)):
                    status_code = str(worksheet.cell(row=row_idx, column=6).value)
                    es_critico = any(err in status_code for err in ["504", "5001", "13", "500"])
                    
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_num)
                        cell.font = font_body
                        cell.border = border_thin
                        cell.alignment = Alignment(vertical="center")
                        
                        if es_critico:
                            cell.fill = fill_error_critico
                
                # 5. Auto-fit column widths dynamically
                for column_cells in worksheet.columns:
                    col_letter = get_column_letter(column_cells[0].column)
                    max_len = 0
                    for cell in column_cells:
                        if cell.row > 4 and cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 45)
                
                worksheet.row_dimensions[5].height = 28

            print(f"[✅] Formateo corporativo premium finalizado con éxito.")
            return self.ruta_salida
        except Exception as e:
            print(f"[-] Error al aplicar diseño al archivo Excel: {e}")
            return None
